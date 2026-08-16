#!/usr/bin/env python3
"""
build_json.py — Hausentscheider-Generator
==========================================
Liest Excel "Der_Entscheider_Testsystem_v2.0_neutral.xlsx" und erzeugt:
  - daten/parameter.json     (Modellannahmen, Block 1-7 nach Excel-Tab "Parameter")
  - daten/preishistorie.json (Marktstaende, Excel-Tab "Preishistorie")

Die Defined Names der Excel sind die Single Source of Truth.
Alle Web-Werkzeuge (rechner.html, objekte/index.html, preishistorie.html,
Mini-Rechner) ziehen ausschliesslich aus den beiden JSON-Dateien.

Aufruf:
    python build_json.py [--xlsx PFAD] [--out daten/]

Standardpfade ergeben sich relativ zum Skript-Speicherort und zum Repo-Layout
(siehe README am Ende dieser Datei).

Idempotent: laeuft beliebig oft, ueberschreibt JSON-Dateien deterministisch.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl fehlt. Installation: pip install openpyxl")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cellref_to_rowcol(ref: str) -> tuple[str, int, int]:
    """'Parameter!$C$8' -> ('Parameter', 8, 3). Quotierte Tab-Namen werden gestripped."""
    sheet, addr = ref.split("!", 1)
    sheet = sheet.strip().strip("'")
    addr = addr.replace("$", "")
    m = re.match(r"([A-Z]+)(\d+)", addr)
    col_letters, row = m.group(1), int(m.group(2))
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return sheet, row, col


def _read_defined_value(wb, name: str) -> Any:
    """Liest den Wert eines Defined Names. Gibt None bei Fehlen."""
    if name not in wb.defined_names:
        return None
    ref = wb.defined_names[name].value
    if not ref or "!" not in ref:
        return None
    sheet, row, col = _cellref_to_rowcol(ref)
    return wb[sheet].cell(row=row, column=col).value


def _row(ws, row_idx: int, max_col: int = 8) -> list[Any]:
    return [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]


# ---------------------------------------------------------------------------
# Block-Reader
# ---------------------------------------------------------------------------

def _energiepreis(ws, default_row: int, steig_row: int, label: str) -> dict[str, Any]:
    """Liest einen Energietraeger-Eintrag aus zwei aufeinanderfolgenden Zeilen.

    Spalten im Block 1:  B=Bezeichnung  C=Default  D=Min  E=Max  F=Einheit  G=Quelle
    """
    d_row = _row(ws, default_row)
    s_row = _row(ws, steig_row)
    return {
        "label":       label,
        "default":     d_row[2],   # C
        "min":         d_row[3],   # D
        "max":         d_row[4],   # E
        "einheit":     d_row[5],   # F
        "quelle":      d_row[6],   # G
        "steigerung":  s_row[2],   # C (z.B. 0.03 = 3%/a)
        "steigerung_min": s_row[3],
        "steigerung_max": s_row[4],
        "steigerung_quelle": s_row[6],
        "kontext":     d_row[1],   # B (Volltext-Bezeichnung mit Hinweis)
    }


def block1_energiepreise(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    # Defined Names: PreisGas C8 / SteigGas C9, PreisFW C11 / SteigFW C12, etc.
    return {
        "PreisGas":     _energiepreis(ws, 8,  9,  "Gas (Effektivpreis netto, Bestandstarif)"),
        "PreisFW":      _energiepreis(ws, 11, 12, "Fernwaerme (Erfurt-Effektivpreis netto)"),
        "PreisWP":      _energiepreis(ws, 14, 15, "Waermestrom (WP-Sondertarif)"),
        "PreisOel":     _energiepreis(ws, 17, 18, "Heizoel (Bestand)"),
        "PreisPellets": _energiepreis(ws, 20, 21, "Pellets (Jahresdurchschnittspreis)"),
    }


def block2_rahmen(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    # BAFA-Stapel + gesetzliche Rahmen + CO2 + Gas-Pfade
    keys_with_meta = [
        ("BonusGrund_default",  28, "BAFA Grundfoerderung WP/Hybrid/Pellets/FW"),
        ("BonusKlima_default",  29, "BAFA Klima-Geschwindigkeitsbonus"),
        ("BonusEink_default",   30, "BAFA Einkommensbonus"),
        ("BonusEff_default",    31, "BAFA Effizienzbonus"),
        ("FoerderDeckel",       32, "BAFA Foerderdeckel relativ"),
        ("FoerderDeckelAbs",    33, "BAFA Foerderdeckel absolut"),
        ("BGB559_alt",          35, "Hinweisreferenz - operativ siehe Block 6 BGB559Korr"),
        ("Kalkzins",            36, "Kalkulationszins (VDI 2067)"),
        ("Zeitraum",            37, "Betrachtungszeitraum (Jahre)"),
        ("CO2Pfad",             39, "CO2-Preis-Pfad (niedrig/mittel/hoch/Custom)"),
        ("CO2PreisCustom_def",  40, "CO2-Preis Custom (EUR/Tonne)"),
        ("Gruengas",            41, "Gruengasquote im Gasmix"),
        ("Biomethan",           42, "Biomethan-Aufschlag (ct/kWh)"),
    ]
    out = {}
    for key, row, label in keys_with_meta:
        d = _row(ws, row)
        out[key] = {
            "label":   label,
            "default": d[2],
            "min":     d[3],
            "max":     d[4],
            "einheit": d[5],
            "quelle":  d[6],
        }
    return out


def block3_technik(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    keys_with_meta = [
        ("JAZ",           49, "JAZ Waermepumpe (Bestand teilsaniert)"),
        ("HybridAnteilWP", 50, "Hybrid-Anteil WP"),
        ("WGGasBW",       51, "Wirkungsgrad Gas-Brennwert"),
        ("WGPellets",     52, "Wirkungsgrad Pellets"),
        ("WartGas",       54, "Wartung Gas-Brennwert"),
        ("WartHybrid",    55, "Wartung Hybrid"),
        ("WartWP",        56, "Wartung Waermepumpe"),
        ("WartFW",        57, "Wartung Fernwaerme"),
        ("WartPellets",   58, "Wartung Pellets"),
        ("GdWUntergrenze",60, "GdW-Soll Erhaltungsruecklage Untergrenze"),
        ("GdWObergrenze", 61, "GdW-Soll Erhaltungsruecklage Obergrenze"),
        ("WPSchallVoll",  63, "WP-Schallpegel Volllast"),
        ("WPSchallNacht", 64, "WP-Schallpegel Nachtmodus"),
        ("TALaermNacht",  65, "TA-Laerm Nachtgrenzwert Mischgebiet"),
    ]
    out = {}
    for key, row, label in keys_with_meta:
        d = _row(ws, row)
        out[key] = {
            "label":   label,
            "default": d[2],
            "min":     d[3],
            "max":     d[4],
            "einheit": d[5],
            "quelle":  d[6],
        }
    return out


def block4_plausi(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    keys_with_meta = [
        ("PelletsMaxWE",        72, "Pellets - max. WE fuer MFH-Sondermarkt"),
        ("PelletsInnenstadtAus",73, "Pellets - Innenstadt-Filter aus"),
        ("OelInvestZulaessig",  75, "Oel als Investitionsoption zulaessig"),
        ("SanRedMax",           77, "Sanierungsregler max. Reduktion"),
    ]
    out = {}
    for key, row, label in keys_with_meta:
        d = _row(ws, row)
        out[key] = {
            "label":   label,
            "default": d[2],
            "min":     d[3],
            "max":     d[4],
            "einheit": d[5],
            "quelle":  d[6],
        }
    return out


def block5_gebaeudedefaults(wb) -> dict[str, Any]:
    """EFH-/MFH-Defaults aus Z124-Z142.
    Spalten: B=Parameter  C=DN_EFH  D=DN_MFH  E=Wert_EFH  F=Wert_MFH  G=Einheit"""
    ws = wb["Parameter"]
    rows = list(range(124, 143))
    efh: dict[str, Any] = {}
    mfh: dict[str, Any] = {}
    meta: dict[str, dict[str, Any]] = {}
    for r in rows:
        row = _row(ws, r)
        label    = row[1]
        dn_efh   = row[2]
        dn_mfh   = row[3]
        val_efh  = row[4]
        val_mfh  = row[5]
        einheit  = row[6]
        if not dn_efh:
            continue
        # Schluesselname ohne EFH_/MFH_-Praefix (= Excel-Stamm, z.B. "Wohnflaeche")
        stem = str(dn_efh).removeprefix("EFH_")
        efh[stem] = val_efh
        mfh[stem] = val_mfh
        meta[stem] = {"label": label, "einheit": einheit,
                      "dn_efh": dn_efh, "dn_mfh": dn_mfh}
    return {"EFH": efh, "MFH": mfh, "_meta": meta}


def block6_vermieter(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    keys_with_meta = [
        ("BGB559Korr",     89, "Paragraf 559 BGB Modernisierungsumlage (gesetzlich seit 2019)"),
        ("Hf559",          90, "Paragraf 559 Haertefall-Abschlag"),
        ("Dauer559",       91, "Paragraf 559 Wirkungsdauer (Jahre)"),
        ("MietspiegelEff", 92, "Mietspiegel-Effekt pro Klassen-Sprung"),
        ("AfA7b",          93, "Paragraf 7b EStG Sonder-AfA"),
        ("AfA7bDauer",     94, "Paragraf 7b EStG Anwendungsdauer"),
        ("KfWZuschuss",    95, "KfW-Tilgungszuschuss EH 55"),
        ("MarktDC",        96, "Marktwert-Premium D->C"),
        ("MarktDB",        97, "Marktwert-Premium D->B"),
        ("EPBDAbschlag",   98, "EPBD-Verkehrswert-Abschlag (Klasse F-H)"),
        ("MietausfQ",      99, "Mietausfall durch Heizkostenanstieg"),
        ("BauMM",         100, "Bauprozess-Mietminderung"),
        ("KaltMiete",     101, "Mittlere Kaltmiete (Beispiel-MFH)"),
        ("Verkehrswert",  102, "Verkehrswert Beispiel-MFH (Schaetzung)"),
        ("KlassenSprung", 103, "Effizienzklassen-Sprung-Annahme"),
        ("Wirksam_VM",    104, "Wirksam Vermieter-Bilanz (intern, dynamisch aus Nutzungsart)"),
    ]
    out = {}
    for key, row, label in keys_with_meta:
        d = _row(ws, row)
        out[key] = {
            "label":   label,
            "default": d[2],
            "min":     d[3],
            "max":     d[4],
            "einheit": d[5],
            "quelle":  d[6],
        }
    return out


def block7_lebensdauer(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    keys_with_meta = [
        ("L_GasBW",   112, "Gas-Brennwert"),
        ("L_Hybrid",  113, "Hybrid (Gas+WP)"),
        ("L_WP",      114, "Waermepumpe"),
        ("L_FW",      115, "Fernwaerme"),
        ("L_Pellets", 116, "Pellets"),
        ("L_Oel",     117, "Oel (Bestand)"),
    ]
    out = {}
    for key, row, label in keys_with_meta:
        d = _row(ws, row)
        out[key] = {
            "label":   label,
            "default": d[2],
            "min":     d[3],
            "max":     d[4],
            "einheit": d[5],
            "quelle":  d[6],
        }
    return out


def stand_und_pflege(wb) -> dict[str, Any]:
    ws = wb["Parameter"]
    return {
        "letzte_aktualisierung": _stand_to_iso(ws.cell(row=80, column=3).value),
        "naechste_pruefung":     _stand_to_iso(ws.cell(row=81, column=3).value),
        "bearbeiter":            ws.cell(row=82, column=3).value,
    }


def _stand_to_iso(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return str(v)


# ---------------------------------------------------------------------------
# Preishistorie
# ---------------------------------------------------------------------------

MONAT_DE_TO_NUM = {
    "januar": "01", "februar": "02", "maerz": "03", "marz": "03", "märz": "03",
    "april": "04", "mai": "05", "juni": "06", "juli": "07",
    "august": "08", "september": "09", "oktober": "10",
    "november": "11", "dezember": "12",
}


def _monatlabel_to_iso(label: str) -> str | None:
    if not label:
        return None
    parts = label.strip().lower().split()
    if len(parts) != 2:
        return label
    monat, jahr = parts
    nr = MONAT_DE_TO_NUM.get(monat)
    if not nr or not jahr.isdigit():
        return label
    return f"{jahr}-{nr}"


def preishistorie(wb) -> dict[str, Any]:
    ws = wb["Preishistorie"]
    # Header: Z5 (Bezeichnungen), Z6 (Einheiten-Subline)
    # Daten ab Z7 absteigend (Mai oben, Januar unten)
    monate: dict[str, Any] = {}
    aktuell_iso: str | None = None
    for r in range(7, 30):
        row = _row(ws, r)
        if not row[0]:
            break
        label = str(row[0]).strip()
        iso = _monatlabel_to_iso(label)
        is_aktuell = (r == 7)
        if is_aktuell:
            aktuell_iso = iso
        eintrag = {
            "label":      label,
            "aktuell":    is_aktuell,
            "gas":        row[1],
            "fernwaerme": row[2],
            "heizoel":    row[3],
            "wp_strom":   row[4],
            "pellets":    row[5],
            "einheit":    row[6],
            "kontext":    row[7],
        }
        monate[iso] = eintrag

    # Methodische Trennung + Fernwaerme-Methodik + Update-Workflow
    methodik_block: list[str] = []
    for r in range(13, 24):
        v = ws.cell(row=r, column=1).value
        if v:
            methodik_block.append(str(v))

    update_workflow: list[str] = []
    for r in range(25, 31):
        v = ws.cell(row=r, column=1).value
        if v:
            update_workflow.append(str(v))

    return {
        "version":         "1.0",
        "stand":           date.today().isoformat(),
        "aktueller_monat": aktuell_iso,
        "spalten_einheit": {
            "gas":        "ct/kWh brutto Verbraucher-Durchschnitt",
            "fernwaerme": "ct/kWh Erfurt-Effektivpreis (tarifbasiert, nicht monatsvolatil)",
            "heizoel":    "ct/kWh brutto Verbraucher-Durchschnitt",
            "wp_strom":   "ct/kWh brutto Verbraucher-Durchschnitt",
            "pellets":    "ct/kWh brutto Verbraucher-Durchschnitt",
        },
        "monate":           monate,
        "methodik":         methodik_block,
        "update_workflow":  update_workflow,
    }


# ---------------------------------------------------------------------------
# Defined-Names-Index (Bruecke Excel <-> JSON)
# ---------------------------------------------------------------------------

def defined_names_index(wb) -> dict[str, str]:
    """Komplette Defined-Names-Tabelle: Name -> Zellreferenz.
    Wird in parameter.json als Anhang mitgeschrieben - dient als Glasbox
    fuer Pruefer und als Bruecke fuer C2-Implementation."""
    out: dict[str, str] = {}
    for name in sorted(wb.defined_names):
        out[name] = wb.defined_names[name].value
    return out


# ---------------------------------------------------------------------------
# Hauptfunktion
# ---------------------------------------------------------------------------

def _pruefe_ueberschreiben(ziel: Path, erzeugt: dict, xlsx_name: str) -> list[str]:
    """Gruende, aus denen 'ziel' nicht ueberschrieben werden darf.

    Zwei Bedingungen, weil eine allein nicht reicht:

    1. Handgepflegte Top-Level-Bloecke, die dieses Skript nicht erzeugt. Sie
       waeren nach dem Schreiben weg. Betroffen sind heute 'foerdersaetze' und
       'foerderstand' in parameter.json sowie 'korrekturen' und
       'methodik_traeger' in preishistorie.json.

    2. Andere Quell-Excel als die, aus der die vorhandene Datei stammt. Das
       faengt den Fall, den Bedingung 1 nicht sieht: Ruecksetzungen INNERHALB
       bekannter Bloecke. Ein Lauf mit einem aelteren Stand stellte zum Beispiel
       PreisGas von 10,2 ct brutto auf 8,57 ct netto zurueck und die fuenf
       Wartungsquoten auf ihre Werte vor dem Angleich vom 05.05.2026 — ohne
       dass die Seite sichtbar bricht.
    """
    if not ziel.exists():
        return []
    try:
        with ziel.open(encoding="utf-8") as f:
            vorhanden = json.load(f)
    except (OSError, ValueError) as err:
        return [f"{ziel.name}: vorhandene Datei nicht lesbar ({err}). Abbruch aus Vorsicht."]
    if not isinstance(vorhanden, dict):
        return [f"{ziel.name}: unerwartete Struktur, kein Objekt. Abbruch aus Vorsicht."]

    gruende: list[str] = []

    verloren = sorted(k for k in vorhanden if k not in erzeugt)
    if verloren:
        gruende.append(
            f"{ziel.name}: {len(verloren)} handgepflegte(r) Block/Bloecke gingen verloren "
            f"({', '.join(verloren)}). Dieses Skript erzeugt sie nicht."
        )

    alte_quelle = vorhanden.get("quelle")
    if alte_quelle and alte_quelle != xlsx_name:
        gruende.append(
            f"{ziel.name}: andere Quelle. Vorhanden ist der Stand aus '{alte_quelle}', "
            f"uebergeben wurde '{xlsx_name}'. Werte koennen innerhalb bekannter Bloecke "
            f"zurueckgesetzt werden, ohne dass es auffaellt."
        )

    return gruende


def build(xlsx_path: Path, out_dir: Path, datenverlust_in_kauf_nehmen: bool = False) -> tuple[Path, Path]:
    if not xlsx_path.exists():
        sys.exit(f"Excel-Quelle nicht gefunden: {xlsx_path}")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    parameter = {
        "version":  "2.0",
        "stand":    date.today().isoformat(),
        "quelle":   xlsx_path.name,
        "tabs":     wb.sheetnames,
        "anzahl_defined_names": len(wb.defined_names),
        "block1_energiepreise":   block1_energiepreise(wb),
        "block2_rahmen":          block2_rahmen(wb),
        "block3_technik":         block3_technik(wb),
        "block4_plausi":          block4_plausi(wb),
        "block5_gebaeudedefaults": block5_gebaeudedefaults(wb),
        "block6_vermieter":       block6_vermieter(wb),
        "block7_lebensdauer":     block7_lebensdauer(wb),
        "stand_und_pflege":       stand_und_pflege(wb),
        "_defined_names_index":   defined_names_index(wb),
    }

    preise = preishistorie(wb)
    preise["quelle"] = xlsx_path.name

    out_dir.mkdir(parents=True, exist_ok=True)
    p_param = out_dir / "parameter.json"
    p_preis = out_dir / "preishistorie.json"

    gruende = (_pruefe_ueberschreiben(p_param, parameter, xlsx_path.name)
               + _pruefe_ueberschreiben(p_preis, preise, xlsx_path.name))
    if gruende:
        kopf = "ABBRUCH — Schreiben wuerde gepflegte Daten verlieren:"
        text = "\n".join("  - " + g for g in gruende)
        if not datenverlust_in_kauf_nehmen:
            sys.exit(
                f"{kopf}\n{text}\n\n"
                "Nichts geschrieben. Sicherung: daten/_sicherung/.\n"
                "Wenn das gewollt ist: erneut mit --datenverlust-in-kauf-nehmen aufrufen."
            )
        print(f"WARNUNG — uebergangen per --datenverlust-in-kauf-nehmen:\n{text}")

    with p_param.open("w", encoding="utf-8") as f:
        json.dump(parameter, f, ensure_ascii=False, indent=2, default=str)
    with p_preis.open("w", encoding="utf-8") as f:
        json.dump(preise, f, ensure_ascii=False, indent=2, default=str)

    return p_param, p_preis


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path(__file__).resolve().parent.parent /
                "02_Finalisierung_Pruefung_Benchmark" /
                "Der_Entscheider_Testsystem_v2.0_neutral.xlsx",
        help="Pfad zur Excel-Quelle (Default: Repo-relativ)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent / "daten",
        help="Zielverzeichnis fuer parameter.json + preishistorie.json",
    )
    parser.add_argument(
        "--datenverlust-in-kauf-nehmen",
        action="store_true",
        help="Schreibt auch dann, wenn handgepflegte Bloecke verloren gehen oder "
             "die Quell-Excel nicht zu der passt, aus der die vorhandene JSON stammt. "
             "Nur nach Sicherung und mit Vorher-nachher-Vergleich verwenden.",
    )
    args = parser.parse_args()

    p_param, p_preis = build(args.xlsx, args.out, args.datenverlust_in_kauf_nehmen)
    print(f"OK   {p_param}")
    print(f"OK   {p_preis}")


if __name__ == "__main__":
    main()


# ---------------------------------------------------------------------------
# README
# ---------------------------------------------------------------------------
# Repo-Layout (Annahme):
#   hausentscheider/
#   ├── build_json.py            <- DIESES SKRIPT
#   ├── daten/
#   │   ├── parameter.json       <- generiert
#   │   └── preishistorie.json   <- generiert
#   ├── rechner.html
#   ├── index.html
#   └── ...
#
# Excel-Quelle liegt eine Ebene hoeher (nicht im Repo):
#   <Workspace>/02_Finalisierung_Pruefung_Benchmark/Der_Entscheider_Testsystem_v2.0_neutral.xlsx
#
# Update-Workflow (monatlich):
#   1. Excel pflegen (Tab Preishistorie - neuen Monat oben einfuegen)
#   2. python build_json.py
#   3. git diff daten/ pruefen - vorher/nachher aller Bloecke ansehen
#   4. git commit -am "Daten Mai 2026" && git push
#   5. Netlify deployt automatisch
#
# ACHTUNG - dieses Skript schreibt parameter.json und preishistorie.json
# VOLLSTAENDIG neu. Handgepflegte Bloecke, die es nicht erzeugt, waeren weg:
#   parameter.json      foerdersaetze, foerderstand
#   preishistorie.json  korrekturen, methodik_traeger
# Ebenso koennen Werte INNERHALB bekannter Bloecke zurueckfallen, wenn eine
# aeltere Excel uebergeben wird (z. B. PreisGas 10,2 ct brutto -> 8,57 netto,
# Wartungsquoten auf den Stand vor dem Angleich vom 05.05.2026).
#
# Deshalb bricht _pruefe_ueberschreiben() in beiden Faellen ab und schreibt
# nichts. Bewusstes Uebergehen: --datenverlust-in-kauf-nehmen, aber nur nach
# Sicherung (daten/_sicherung/) und mit Vorher-nachher-Vergleich.
#
# Stand 16.08.2026: parameter.json nennt als Quelle v2.1_neutral_OUTPUT.xlsx
# (Stand 2026-05-02). Diese Datei ist im Repo nicht auffindbar; vorhanden ist
# nur Der_Entscheider_Testsystem_v2.0_neutral.xlsx unter dokumente/untracked/.
# Ein Lauf damit waere ein Rueckschritt, kein Update.
#
# Sicherheitsnetz: Falls Excel korrupt wird (OneDrive-Sync-Problem),
# die unbeschaedigte Variante als Quelle nutzen oder aus v1.6_Theaterstrasse
# rekonstruieren.
